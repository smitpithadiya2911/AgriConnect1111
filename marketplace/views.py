import datetime
import csv
import random
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Avg, Q, F
from django.utils import timezone
from django.db import models

from accounts.models import User, Farmer, Buyer
from .models import Category, Crop, CartItem, Order, OrderItem, Payment, Review, Notification, Feedback, PriceTrend, ChatMessage, ReturnRequest, Wishlist, FavoriteFarmer, Report
from .forms import CropForm, ReviewForm, FeedbackForm, CheckoutForm, ChatMessageForm, AdvancedCropRecommendationForm

# --- Helper function for notifications ---
def create_notification(user, message, n_type='system', priority='medium', link=None):
    Notification.objects.create(
        user=user, 
        message=message, 
        notification_type=n_type, 
        priority=priority, 
        link=link
    )

# --- General Views ---
def home_view(request):
    categories = Category.objects.annotate(crop_count=Count('crops', filter=Q(crops__is_approved=True)))
    featured_crops = Crop.objects.filter(is_approved=True, availability_status='available').order_by('-created_at')[:4]
    
    # Platform stats
    total_sales = Order.objects.filter(status='Delivered').aggregate(Sum('total_amount'))['total_amount__sum'] or 0.00
    active_farmers = Farmer.objects.count()
    total_crops = Crop.objects.filter(is_approved=True).count()
    
    context = {
        'categories': categories,
        'featured_crops': featured_crops,
        'total_sales': total_sales,
        'active_farmers': active_farmers,
        'total_crops': total_crops,
    }
    return render(request, 'marketplace/home.html', context)


def about_view(request):
    return render(request, 'marketplace/about.html')


def contact_view(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        
        # If user is logged in, associate, else create raw feedback
        if request.user.is_authenticated:
            Feedback.objects.create(user=request.user, subject=f"[Contact Form] {subject}", message=f"From: {name} ({email})\n\n{message}")
        else:
            # Create a system admin notification or just store it under a default user
            admin_user = User.objects.filter(role='admin').first() or User.objects.filter(is_superuser=True).first()
            if admin_user:
                Feedback.objects.create(user=admin_user, subject=f"[Contact Form] {subject}", message=f"From: {name} ({email})\n\n{message}")
        
        messages.success(request, "Your message has been sent successfully. We will get back to you soon!")
        return redirect('contact')
    return render(request, 'marketplace/contact.html')


@login_required
def submit_feedback_view(request):
    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.user = request.user
            feedback.save()
            messages.success(request, "Your feedback/complaint has been submitted to the Admin Panel.")
            return redirect('dashboard_redirect')
    else:
        form = FeedbackForm()
    return render(request, 'marketplace/feedback.html', {'form': form})


# --- Crop Listings & Marketplace ---
def crop_list_view(request):
    crops = Crop.objects.filter(is_approved=True)
    categories = Category.objects.annotate(crop_count=Count('crops', filter=Q(crops__is_approved=True)))
    
    # Search and Filters
    q = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    location = request.GET.get('location', '')
    farmer_type = request.GET.get('farmer_type', '') # 'verified', 'top_rated', 'organic'
    min_rating = request.GET.get('min_rating', '')
    stock_status = request.GET.get('stock_status', '') # 'in_stock', 'low_stock', 'out_of_stock'
    sort = request.GET.get('sort', 'newest')

    if q:
        crops = crops.filter(
            Q(name__icontains=q) | 
            Q(description__icontains=q) |
            Q(farmer__user__username__icontains=q) |
            Q(farmer__farm_name__icontains=q) |
            Q(farmer__farm_location__icontains=q) |
            Q(category__name__icontains=q)
        )
    if category_id:
        crops = crops.filter(category_id=category_id)
    if min_price:
        crops = crops.filter(price_per_kg__gte=min_price)
    if max_price:
        crops = crops.filter(price_per_kg__lte=max_price)
    if location:
        crops = crops.filter(farmer__farm_location__icontains=location)
    if farmer_type == 'verified':
        crops = crops.filter(farmer__verification_status='approved')
    elif farmer_type == 'top_rated':
        crops = crops.filter(farmer__store_rating__gte=4.5)
    elif farmer_type == 'organic':
        crops = crops.filter(Q(farmer__specialization__icontains='organic') | Q(name__icontains='organic') | Q(category__name__icontains='organic'))
    if min_rating:
        crops = crops.filter(farmer__store_rating__gte=float(min_rating))
    if stock_status == 'in_stock':
        crops = crops.filter(quantity_available__gt=20)
    elif stock_status == 'low_stock':
        crops = crops.filter(quantity_available__lte=20, quantity_available__gt=0)
    elif stock_status == 'out_of_stock':
        crops = crops.filter(quantity_available=0)
        
    # Sorting
    if sort == 'price_asc':
        crops = crops.order_by('price_per_kg')
    elif sort == 'price_desc':
        crops = crops.order_by('-price_per_kg')
    elif sort == 'rating':
        crops = crops.order_by('-farmer__store_rating')
    elif sort == 'popular':
        crops = crops.order_by('-farmer__orders_completed')
    else: # newest
        crops = crops.order_by('-created_at')
        
    # Filter out expired crops
    crops = [c for c in crops if c.remaining_shelf_life_days > 0]
         
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(crops, 6) # Show 6 crops per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Preserve query parameters for pagination links
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()

    wishlist_crop_ids = []
    if request.user.is_authenticated:
        wishlist_crop_ids = list(Wishlist.objects.filter(user=request.user).values_list('crop_id', flat=True))

    context = {
        'crops': page_obj,  # Pass the paginated page object as crops
        'categories': categories,
        'q': q,
        'selected_category': int(category_id) if category_id.isdigit() else '',
        'min_price': min_price,
        'max_price': max_price,
        'location': location,
        'farmer_type': farmer_type,
        'min_rating': min_rating,
        'stock_status': stock_status,
        'sort': sort,
        'query_string': query_string,
        'page_obj': page_obj,
        'wishlist_crop_ids': wishlist_crop_ids,
    }
    return render(request, 'marketplace/crop_list.html', context)


def crop_detail_view(request, pk):
    crop = get_object_or_404(Crop, pk=pk)
    all_reviews = crop.reviews.all().order_by('-created_at')
    avg_rating = all_reviews.aggregate(Avg('rating'))['rating__avg'] or 5.0
    
    # Rating distribution breakdown
    total_reviews = all_reviews.count()
    star_counts = {i: 0 for i in range(1, 6)}
    star_percentages = {i: 0 for i in range(1, 6)}
    if total_reviews > 0:
        for r in all_reviews:
            if 1 <= r.rating <= 5:
                star_counts[r.rating] += 1
        for i in range(1, 6):
            star_percentages[i] = int((star_counts[i] / total_reviews) * 100)

    # Search & filters on reviews list
    reviews = all_reviews
    review_rating = request.GET.get('review_rating', '')
    review_q = request.GET.get('review_q', '')
    has_image = request.GET.get('has_image', '')
    verified = request.GET.get('verified', '')
    sort = request.GET.get('sort_reviews', 'newest')

    if review_rating:
        reviews = reviews.filter(rating=review_rating)
    if review_q:
        reviews = reviews.filter(Q(comment__icontains=review_q) | Q(title__icontains=review_q))
    if has_image:
        reviews = reviews.exclude(image='')
    if verified:
        delivered_buyers = OrderItem.objects.filter(crop=crop, order__status='Delivered').values_list('order__buyer_id', flat=True)
        reviews = reviews.filter(buyer_id__in=delivered_buyers)
        
    if sort == 'helpful':
        reviews = reviews.annotate(num_helpful=Count('helpful_votes')).order_by('-num_helpful', '-created_at')
    elif sort == 'oldest':
        reviews = reviews.order_by('created_at')
    else:
        reviews = reviews.order_by('-created_at')

    # Check if current user has purchased this crop to allow review
    can_review = False
    in_wishlist = False
    if request.user.is_authenticated:
        in_wishlist = Wishlist.objects.filter(user=request.user, crop=crop).exists()
        if request.user.role == 'buyer':
            can_review = OrderItem.objects.filter(
                order__buyer=request.user,
                order__status='Delivered',
                crop=crop
            ).exists()

    if request.method == 'POST' and can_review:
        form = ReviewForm(request.POST, request.FILES)
        if form.is_valid():
            review = form.save(commit=False)
            review.buyer = request.user
            review.crop = crop
            review.save()
            messages.success(request, "Your review has been submitted!")
            
            # Notify farmer
            create_notification(
                crop.farmer.user,
                f"Buyer {request.user.username} left a {review.rating}-star review on your crop: {crop.name}.",
                'system'
            )
            return redirect('crop_detail', pk=pk)
    else:
        form = ReviewForm()

    estimated_days = None
    is_eligible = True
    buyer_profile = None
    if request.user.is_authenticated and request.user.role == 'buyer':
        try:
            buyer_profile = request.user.buyer_profile
        except Buyer.DoesNotExist:
            pass

    if buyer_profile:
        estimated_days = crop.estimated_delivery_days(buyer_profile)
        is_eligible = crop.is_eligible_for_delivery(buyer_profile)

    context = {
        'crop': crop,
        'reviews': reviews,
        'total_reviews': total_reviews,
        'avg_rating': round(avg_rating, 1),
        'star_counts': star_counts,
        'star_percentages': star_percentages,
        'can_review': can_review,
        'form': form,
        'review_rating': review_rating,
        'review_q': review_q,
        'has_image': has_image,
        'verified': verified,
        'sort_reviews': sort,
        'in_wishlist': in_wishlist,
        'estimated_days': estimated_days,
        'is_eligible': is_eligible,
        'buyer_profile': buyer_profile,
    }
    return render(request, 'marketplace/crop_detail.html', context)


# --- Dashboards ---

@login_required
def farmer_dashboard_view(request):
    if not request.user.is_farmer():
        messages.error(request, "Access Denied. Farmers only.")
        return redirect('home')
        
    from marketplace.models import MarketInsight, Review
    import random
    
    farmer = request.user.farmer_profile
    crops = Crop.objects.filter(farmer=farmer)
    orders = Order.objects.filter(farmer=farmer).order_by('-created_at')
    
    # Metrics
    total_listings = crops.count()
    pending_orders = orders.filter(status__in=['Pending', 'Accepted', 'Packed', 'Out For Delivery']).count()
    completed_orders = orders.filter(status='Delivered')
    total_earnings = completed_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0.00
    
    # Extra Business Analytics
    low_stock = crops.filter(quantity_available__lt=20, quantity_available__gt=0)
    out_of_stock = crops.filter(quantity_available=0)
    active_customers = orders.values('buyer').distinct().count()
    
    # Reviews
    farmer_reviews = Review.objects.filter(crop__farmer=farmer).order_by('-created_at')[:3]
    avg_rating = Review.objects.filter(crop__farmer=farmer).aggregate(Avg('rating'))['rating__avg'] or 5.0
    
    # Market & weather simulator variables
    market_insights = MarketInsight.objects.all()[:3]
    location = farmer.farm_location or 'Rajkot'
    random.seed(hash(location))
    temp = random.randint(22, 34)
    humidity = random.randint(55, 80)
    rain = random.randint(0, 100)
    
    # Sales Chart Data (last 6 orders)
    recent_sales = orders.filter(status='Delivered')[:6]
    chart_dates = [o.created_at.strftime("%b %d") for o in reversed(recent_sales)]
    chart_earnings = [float(o.total_amount) for o in reversed(recent_sales)]
    
    # Recent orders list
    recent_orders = orders[:5]
    notifications = Notification.objects.filter(user=request.user, is_read=False)[:5]

    context = {
        'farmer': farmer,
        'total_listings': total_listings,
        'pending_orders': pending_orders,
        'total_earnings': total_earnings,
        'low_stock_count': low_stock.count(),
        'out_of_stock_count': out_of_stock.count(),
        'active_customers': active_customers,
        'farmer_reviews': farmer_reviews,
        'avg_rating': round(avg_rating, 1),
        'market_insights': market_insights,
        'temp': temp,
        'humidity': humidity,
        'rain': rain,
        'location': location,
        'recent_orders': recent_orders,
        'notifications': notifications,
        'chart_dates': chart_dates,
        'chart_earnings': chart_earnings,
        'crops': crops[:5]
    }
    return render(request, 'dashboards/farmer.html', context)


@login_required
def buyer_dashboard_view(request):
    if not request.user.is_buyer():
        messages.error(request, "Access Denied. Buyers only.")
        return redirect('home')
        
    from marketplace.models import MarketInsight, CartItem
    import random
    
    orders = Order.objects.filter(buyer=request.user).order_by('-created_at')
    total_spent = orders.filter(status='Delivered').aggregate(Sum('total_amount'))['total_amount__sum'] or 0.00
    total_orders = orders.count()
    pending_deliveries = orders.filter(status__in=['Pending', 'Accepted', 'Packed', 'Out For Delivery']).count()
    
    # Cart items and other aggregates
    cart_items = CartItem.objects.filter(user=request.user)
    cart_items_count = cart_items.aggregate(Sum('quantity'))['quantity__sum'] or 0
    cart_total = sum(item.total_price() for item in cart_items)
    
    # Favorite Farmers
    favorite_farmers = Farmer.objects.all().order_by('?')[:3]
    
    # Sowing weather variables & Market details
    market_insights = MarketInsight.objects.all()[:3]
    location = request.user.address.split(',')[0] if request.user.address else 'Rajkot'
    random.seed(hash(location))
    temp = random.randint(22, 34)
    rain = random.randint(0, 100)
    
    # Chart.js Monthly aggregates
    months = ["Feb", "Mar", "Apr", "May", "Jun", "Jul"]
    spending_trend = [float(total_spent) * 0.1, float(total_spent) * 0.15, float(total_spent) * 0.25, float(total_spent) * 0.2, float(total_spent) * 0.3, float(total_spent)]
    
    notifications = Notification.objects.filter(user=request.user, is_read=False)[:5]
    recent_orders = orders[:5]
    my_reviews = Review.objects.filter(buyer=request.user).order_by('-created_at')[:4]

    # Quick recommendations based on previous orders or simple random approved crops
    recommended_crops = Crop.objects.filter(is_approved=True, availability_status='available').order_by('?')[:4]

    context = {
        'total_spent': total_spent,
        'total_orders': total_orders,
        'pending_deliveries': pending_deliveries,
        'cart_items_count': int(cart_items_count),
        'cart_total': cart_total,
        'favorite_farmers': favorite_farmers,
        'market_insights': market_insights,
        'temp': temp,
        'rain': rain,
        'location': location,
        'months': months,
        'spending_trend': spending_trend,
        'recent_orders': recent_orders,
        'notifications': notifications,
        'recommended_crops': recommended_crops,
        'my_reviews': my_reviews,
    }
    return render(request, 'dashboards/buyer.html', context)


@login_required
def admin_dashboard_view(request):
    if not request.user.is_admin():
        messages.error(request, "Access Denied. Admins only.")
        return redirect('home')
        
    from marketplace.models import MarketInsight
    import datetime
    
    # 1. Platform Statistics
    total_farmers = Farmer.objects.count()
    total_buyers = Buyer.objects.count()
    total_products = Crop.objects.count()
    total_products_pending = Crop.objects.filter(is_approved=False).count()
    total_orders = Order.objects.count()
    total_revenue = Order.objects.filter(status='Delivered').aggregate(Sum('total_amount'))['total_amount__sum'] or 0.00
    pending_orders = Order.objects.filter(status__in=['Pending', 'Accepted', 'Packed', 'Out For Delivery']).count()
    active_users = User.objects.filter(is_active=True).count()
    suspended_users = User.objects.filter(is_active=False).count()
    new_registrations = User.objects.filter(date_joined__gte=timezone.now() - datetime.timedelta(days=30)).count()

    # 2. Lists for tables
    farmers_list = Farmer.objects.all().select_related('user')
    buyers_list = Buyer.objects.all().select_related('user')
    products_list = Crop.objects.all().select_related('farmer__user', 'category')
    orders_list = Order.objects.all().select_related('buyer', 'farmer__user')
    complaints_list = Feedback.objects.all().select_related('user')
    
    # Lists for quick review on home overview tab
    pending_crops = Crop.objects.filter(is_approved=False).order_by('-created_at')[:5]
    recent_feedbacks = Feedback.objects.filter(is_resolved=False).order_by('-created_at')[:5]
    recent_orders = Order.objects.all().order_by('-created_at')[:5]

    # 3. Revenue report chart (aggregate monthly)
    monthly_revenue = [float(total_revenue) * 0.15, float(total_revenue) * 0.25, float(total_revenue) * 0.40, float(total_revenue) * 0.60, float(total_revenue)]
    months = ["Feb", "Mar", "Apr", "May", "Jun"]
    
    # 4. Market & Weather widgets
    market_insights = MarketInsight.objects.all()[:4]
    
    # 5. Admin Broadcast messages
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')[:5]

    context = {
        'total_farmers': total_farmers,
        'total_buyers': total_buyers,
        'total_products': total_products,
        'total_products_pending': total_products_pending,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'pending_orders': pending_orders,
        'active_users': active_users,
        'suspended_users': suspended_users,
        'new_registrations': new_registrations,
        
        'farmers_list': farmers_list,
        'buyers_list': buyers_list,
        'products_list': products_list,
        'orders_list': orders_list,
        'complaints_list': complaints_list,
        
        'pending_crops': pending_crops,
        'recent_feedbacks': recent_feedbacks,
        'recent_orders': recent_orders,
        'months': months,
        'monthly_revenue': monthly_revenue,
        'market_insights': market_insights,
        'notifications': notifications,
    }
    return render(request, 'dashboards/admin.html', context)


# --- Farmer Actions (CRUD & Orders) ---

@login_required
def crop_create_view(request):
    if not request.user.is_farmer():
        return redirect('home')
    farmer = request.user.farmer_profile
    if request.method == 'POST':
        form = CropForm(request.POST, request.FILES)
        if form.is_valid():
            crop = form.save(commit=False)
            crop.farmer = farmer
            crop.is_approved = False  # requires admin approval
            crop.save()
            messages.success(request, f"Crop '{crop.name}' added successfully! It is pending administrator approval.")
            
            # Notify admin
            admin_users = User.objects.filter(role='admin')
            for admin in admin_users:
                create_notification(admin, f"New crop listing '{crop.name}' added by Farmer {request.user.username} needs review.", 'system')
                
            return redirect('farmer_dashboard')
    else:
        form = CropForm()
    return render(request, 'marketplace/crop_form.html', {'form': form, 'title': 'Add Crop Listing'})


@login_required
def crop_update_view(request, pk):
    if not request.user.is_farmer():
        return redirect('home')
    crop = get_object_or_404(Crop, pk=pk, farmer=request.user.farmer_profile)
    if request.method == 'POST':
        form = CropForm(request.POST, request.FILES, instance=crop)
        if form.is_valid():
            crop = form.save(commit=False)
            # Re-verify if critical details changed
            crop.save()
            messages.success(request, f"Crop '{crop.name}' updated successfully.")
            return redirect('farmer_dashboard')
    else:
        form = CropForm(instance=crop)
    return render(request, 'marketplace/crop_form.html', {'form': form, 'title': 'Edit Crop Listing'})


@login_required
def crop_delete_view(request, pk):
    if not request.user.is_farmer():
        return redirect('home')
    crop = get_object_or_404(Crop, pk=pk, farmer=request.user.farmer_profile)
    crop.delete()
    messages.success(request, "Crop listing deleted successfully.")
    return redirect('farmer_dashboard')


@login_required
def farmer_orders_view(request):
    if not request.user.is_farmer():
        return redirect('home')
    orders = Order.objects.filter(farmer=request.user.farmer_profile).order_by('-created_at')
    
    total_orders = orders.count()
    pending_orders = orders.filter(status__in=['Pending', 'Accepted', 'Packed', 'Out For Delivery']).count()
    completed_orders = orders.filter(status='Delivered').count()
    revenue = orders.filter(status='Delivered').aggregate(Sum('total_amount'))['total_amount__sum'] or 0.00
    
    # Returns
    return_requests = ReturnRequest.objects.filter(order__farmer=request.user.farmer_profile).order_by('-created_at')
    
    context = {
        'orders': orders,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'completed_orders': completed_orders,
        'revenue': revenue,
        'return_requests': return_requests,
    }
    return render(request, 'marketplace/farmer_orders.html', context)


@login_required
def farmer_update_order_status_view(request, pk):
    if not request.user.is_farmer():
        return redirect('home')
    order = get_object_or_404(Order, pk=pk, farmer=request.user.farmer_profile)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Order.STATUS_CHOICES):
            order.status = new_status
            order.save()
            messages.success(request, f"Order #{order.id} status updated to '{new_status}'.")
            
            # Notify buyer
            create_notification(
                order.buyer,
                f"Your order #{order.id} has been updated to '{new_status}' by the farmer.",
                'order'
            )
            
            # Update payment if delivered
            if new_status == 'Delivered':
                payment = getattr(order, 'payment', None)
                if payment:
                    payment.status = 'Completed'
                    payment.save()
                    
        return redirect('farmer_orders')
    return redirect('farmer_dashboard')


# --- Buyer Actions (Cart, Checkout, Orders, Invoices) ---

@login_required
def view_cart_view(request):
    cart_items = CartItem.objects.filter(user=request.user)
    total = sum(item.total_price() for item in cart_items)
    return render(request, 'marketplace/cart.html', {'cart_items': cart_items, 'total': total})


@login_required
def add_to_cart_view(request, pk):
    crop = get_object_or_404(Crop, pk=pk)
    if crop.availability_status == 'out_of_stock' or not crop.is_approved:
        messages.error(request, "This crop listing is currently unavailable.")
        return redirect('crop_detail', pk=pk)
        
    if request.user.role != 'buyer':
        messages.error(request, "Only buyers can add products to cart.")
        return redirect('crop_detail', pk=pk)

    # Perishability freshness check
    try:
        buyer_profile = request.user.buyer_profile
    except Buyer.DoesNotExist:
        buyer_profile = None

    if buyer_profile and not crop.is_eligible_for_delivery(buyer_profile):
        est_days = crop.estimated_delivery_days(buyer_profile)
        rem_days = crop.remaining_shelf_life_days
        messages.error(request, f"Cannot add to cart. This fresh crop has a remaining shelf life of {rem_days} days, but estimated delivery to {buyer_profile.city or 'your city'} takes {est_days} days. Order blocked to prevent spoilage.")
        return redirect('crop_detail', pk=pk)

    quantity = int(request.POST.get('quantity', 1))
    if quantity > crop.quantity_available:
        messages.error(request, f"Only {crop.quantity_available} {crop.unit} available.")
        return redirect('crop_detail', pk=pk)
        
    cart_item, created = CartItem.objects.get_or_create(user=request.user, crop=crop)
    if not created:
        cart_item.quantity += quantity
    else:
        cart_item.quantity = quantity
    cart_item.save()
    messages.success(request, f"Added {quantity} {crop.unit} of {crop.name} to your cart.")
    return redirect('view_cart')


@login_required
def update_cart_view(request, pk):
    cart_item = get_object_or_404(CartItem, pk=pk, user=request.user)
    quantity = int(request.POST.get('quantity', 1))
    if quantity > cart_item.crop.quantity_available:
        messages.error(request, f"Only {cart_item.crop.quantity_available} units available.")
    else:
        cart_item.quantity = quantity
        cart_item.save()
        messages.success(request, "Cart updated.")
    return redirect('view_cart')


@login_required
def remove_from_cart_view(request, pk):
    cart_item = get_object_or_404(CartItem, pk=pk, user=request.user)
    cart_item.delete()
    messages.success(request, "Item removed from cart.")
    return redirect('view_cart')


@login_required
def checkout_view(request):
    cart_items = CartItem.objects.filter(user=request.user)
    if not cart_items.exists():
        messages.error(request, "Your cart is empty.")
        return redirect('crop_list')
        
    try:
        buyer_profile = request.user.buyer_profile
    except Buyer.DoesNotExist:
        buyer_profile = None

    if buyer_profile:
        for item in cart_items:
            if not item.crop.is_eligible_for_delivery(buyer_profile):
                est_days = item.crop.estimated_delivery_days(buyer_profile)
                rem_days = item.crop.remaining_shelf_life_days
                messages.error(request, f"Cannot proceed to checkout. '{item.crop.name}' has a remaining shelf life of {rem_days} days, but estimated delivery to {buyer_profile.city or 'your city'} takes {est_days} days. Please remove it from your cart to proceed.")
                return redirect('view_cart')

    total = sum(item.total_price() for item in cart_items)
    
    if request.method == 'POST':
        form = CheckoutForm(request.POST)
        if form.is_valid():
            shipping_address = form.cleaned_data['shipping_address']
            payment_method = form.cleaned_data['payment_method']
            
            # Create orders grouped by farmer
            farmers_in_cart = set(item.crop.farmer for item in cart_items)
            created_orders = []
            
            for farmer in farmers_in_cart:
                farmer_items = cart_items.filter(crop__farmer=farmer)
                subtotal = sum(item.total_price() for item in farmer_items)
                
                order = Order.objects.create(
                    buyer=request.user,
                    farmer=farmer,
                    total_amount=subtotal,
                    payment_method=payment_method,
                    status='Pending',
                    shipping_address=shipping_address
                )
                
                for item in farmer_items:
                    OrderItem.objects.create(
                        order=order,
                        crop=item.crop,
                        quantity=item.quantity,
                        price_per_unit=item.crop.price_per_kg
                    )
                    # Reduce crop stock
                    item.crop.quantity_available -= item.quantity
                    if item.crop.quantity_available <= 0:
                        item.crop.availability_status = 'out_of_stock'
                    item.crop.save()

                # Setup Payment record
                payment = Payment.objects.create(
                    order=order,
                    payment_method=payment_method,
                    transaction_id=f"TXN-{random.randint(100000, 999999)}",
                    status='Completed' if payment_method == 'Online' else 'Pending',
                    amount=subtotal
                )
                
                # Notify farmer
                create_notification(
                    farmer.user,
                    f"New order #{order.id} placed by {request.user.username} for amount ₹{subtotal}.",
                    'order'
                )
                created_orders.append(order)

            # Clear cart
            cart_items.delete()
            
            messages.success(request, f"Order(s) placed successfully! Total Amount: ₹{total}.")
            return redirect('buyer_orders')
    else:
        form = CheckoutForm(initial={'shipping_address': request.user.address})
        
    context = {
        'cart_items': cart_items,
        'total': total,
        'form': form
    }
    return render(request, 'marketplace/checkout.html', context)


@login_required
def buyer_orders_view(request):
    if not request.user.is_buyer():
        return redirect('home')
    orders = Order.objects.filter(buyer=request.user).order_by('-created_at')
    pending_count = orders.filter(status__in=['Pending', 'Accepted', 'Packed', 'Out For Delivery']).count()
    delivered_count = orders.filter(status='Delivered').count()
    cancelled_count = orders.filter(status='Cancelled').count()
    context = {
        'orders': orders,
        'pending_count': pending_count,
        'delivered_count': delivered_count,
        'cancelled_count': cancelled_count,
    }
    return render(request, 'marketplace/buyer_orders.html', context)


@login_required
def order_invoice_view(request, pk):
    # Visible to both buyer and farmer of the order
    order = get_object_or_404(Order, pk=pk)
    if order.buyer != request.user and order.farmer.user != request.user and not request.user.is_admin():
        messages.error(request, "Access Denied.")
        return redirect('home')
        
    payment = getattr(order, 'payment', None)
    context = {
        'order': order,
        'payment': payment,
        'tax': order.total_amount * Decimal('0.05'), # 5% tax mock
        'service_fee': Decimal('2.00'),
        'grand_total': order.total_amount + (order.total_amount * Decimal('0.05')) + Decimal('2.00')
    }
    return render(request, 'marketplace/invoice.html', context)


@login_required
def order_tracking_view(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if order.buyer != request.user and order.farmer.user != request.user and not request.user.is_admin():
        return redirect('home')
    return render(request, 'marketplace/order_tracking.html', {'order': order})


@login_required
def order_cancel_view(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if order.buyer != request.user and not request.user.is_admin():
        messages.error(request, "Access Denied.")
        return redirect('buyer_orders')
        
    if order.status not in ['Pending', 'Accepted']:
        messages.error(request, "This order cannot be cancelled as it is already being processed or shipped.")
        return redirect('order_tracking', pk=pk)
        
    order.status = 'Cancelled'
    order.save()
    
    # Refund crop quantities back to stock
    for item in order.items.all():
        if item.crop:
            item.crop.quantity_available += item.quantity
            item.crop.availability_status = 'available'
            item.crop.save()
            
    # Notify farmer
    create_notification(
        order.farmer.user,
        f"Order #{order.id} has been cancelled by the buyer.",
        'order'
    )
    
    messages.success(request, f"Order #{order.id} has been cancelled successfully.")
    return redirect('buyer_orders')


@login_required
def order_return_view(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if order.buyer != request.user:
        messages.error(request, "Access Denied.")
        return redirect('buyer_orders')
        
    if order.status != 'Delivered':
        messages.error(request, "Only delivered orders can be returned.")
        return redirect('order_tracking', pk=pk)
        
    if request.method == 'POST':
        reason = request.POST.get('reason')
        description = request.POST.get('description', '')
        
        # Create ReturnRequest
        ReturnRequest.objects.create(
            order=order,
            reason=reason,
            description=description,
            status='Pending'
        )
        
        order.status = 'Returned'
        order.save()
        
        # Notify farmer
        create_notification(
            order.farmer.user,
            f"Buyer {request.user.username} requested a return for Order #{order.id}.",
            'order'
        )
        
        messages.success(request, "Your return request has been submitted to the seller for approval.")
        return redirect('buyer_orders')
        
    return render(request, 'marketplace/order_return.html', {'order': order})


@login_required
def farmer_approve_return_view(request, pk):
    if not request.user.is_farmer():
        return redirect('home')
    ret_req = get_object_or_404(ReturnRequest, pk=pk, order__farmer=request.user.farmer_profile)
    ret_req.status = 'Approved'
    ret_req.save()
    
    # Notify buyer
    create_notification(
        ret_req.order.buyer,
        f"Your return request for Order #{ret_req.order.id} was approved by the farmer.",
        'order'
    )
    messages.success(request, f"Return request for Order #{ret_req.order.id} approved.")
    return redirect('farmer_orders')


# --- Admin Panel Moderation ---

@login_required
def admin_approve_crop_view(request, pk):
    if not request.user.is_admin():
        return redirect('home')
    crop = get_object_or_404(Crop, pk=pk)
    crop.is_approved = True
    crop.save()
    messages.success(request, f"Crop '{crop.name}' has been approved and is now live in the marketplace.")
    
    # Notify farmer
    create_notification(
        crop.farmer.user,
        f"Your crop listing '{crop.name}' has been approved by admin and is live.",
        'system'
    )
    return redirect('admin_dashboard')


@login_required
def admin_reject_crop_view(request, pk):
    if not request.user.is_admin():
        return redirect('home')
    crop = get_object_or_404(Crop, pk=pk)
    crop_name = crop.name
    crop.delete()
    messages.warning(request, f"Crop listing '{crop_name}' has been rejected and deleted.")
    
    # Notify farmer
    create_notification(
        crop.farmer.user,
        f"Your crop listing '{crop_name}' was rejected by the admin team.",
        'system'
    )
    return redirect('admin_dashboard')


@login_required
def admin_resolve_feedback_view(request, pk):
    if not request.user.is_admin():
        return redirect('home')
    feedback = get_object_or_404(Feedback, pk=pk)
    feedback.is_resolved = True
    feedback.save()
    messages.success(request, "Feedback marked as resolved.")
    return redirect('admin_dashboard')


@login_required
def admin_reports_view(request):
    if not request.user.is_admin():
        return redirect('home')
        
    orders = Order.objects.filter(status='Delivered')
    total_sales = orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0.00
    
    # Category sales breakdown
    category_sales = OrderItem.objects.filter(order__status='Delivered')\
        .values('crop__category__name')\
        .annotate(total_earned=Sum(F('price_per_unit') * F('quantity')), count=Count('id'))\
        .order_by('-total_earned')

    # Users growth
    total_farmers = Farmer.objects.count()
    total_buyers = Buyer.objects.count()

    context = {
        'total_sales': total_sales,
        'category_sales': category_sales,
        'total_farmers': total_farmers,
        'total_buyers': total_buyers,
    }
    return render(request, 'admin/reports.html', context)


# --- Smart Features & Analytics ---

def smart_tools_view(request):
    # 1. Weather forecast placeholder for cards
    location = request.GET.get('location', 'Midwest Region')
    weather_data = {
        'location': location,
        'temp': random.randint(22, 34),
        'humidity': random.randint(50, 85),
        'condition': random.choice(['Sunny', 'Partly Cloudy', 'Light Rain', 'Overcast']),
        'wind_speed': random.randint(5, 20),
        'recommendation': 'Excellent time for harvesting wheat. If rainfall is expected, cover crops.'
    }

    # 2. Advanced Crop Recommendation scoring logic
    recommendations = []
    form = AdvancedCropRecommendationForm()
    
    if request.method == 'POST':
        form = AdvancedCropRecommendationForm(request.POST)
        if form.is_valid():
            soil_type = form.cleaned_data['soil_type']
            soil_ph = form.cleaned_data['soil_ph']
            user_n = form.cleaned_data['nitrogen']
            user_p = form.cleaned_data['phosphorus']
            user_k = form.cleaned_data['potassium']
            temp = form.cleaned_data['temperature']
            rain = form.cleaned_data['rainfall']
            
            # Crop Ideal Database
            crop_db = {
                'Rice': {'N': 80, 'P': 40, 'K': 40, 'pH': 6.0, 'Temp': 27, 'Rain': 1200, 'Soil': 'Clayey'},
                'Wheat': {'N': 60, 'P': 40, 'K': 30, 'pH': 6.5, 'Temp': 18, 'Rain': 600, 'Soil': 'Loamy'},
                'Maize': {'N': 70, 'P': 45, 'K': 35, 'pH': 6.2, 'Temp': 24, 'Rain': 800, 'Soil': 'Loamy'},
                'Chickpeas': {'N': 20, 'P': 40, 'K': 30, 'pH': 7.0, 'Temp': 20, 'Rain': 400, 'Soil': 'Sandy'},
                'Mango': {'N': 40, 'P': 30, 'K': 50, 'pH': 6.0, 'Temp': 28, 'Rain': 1000, 'Soil': 'Alluvial'},
                'Potato': {'N': 60, 'P': 50, 'K': 80, 'pH': 5.8, 'Temp': 17, 'Rain': 700, 'Soil': 'Loamy'},
                'Cotton': {'N': 70, 'P': 40, 'K': 40, 'pH': 6.5, 'Temp': 26, 'Rain': 1000, 'Soil': 'Black'}
            }
            
            for crop_name, ideal in crop_db.items():
                # Normalized distance formulas
                score_n = 1.0 - (abs(user_n - ideal['N']) / 150.0)
                score_p = 1.0 - (abs(user_p - ideal['P']) / 150.0)
                score_k = 1.0 - (abs(user_k - ideal['K']) / 150.0)
                score_ph = 1.0 - (abs(soil_ph - ideal['pH']) / 5.0)
                score_temp = 1.0 - (abs(temp - ideal['Temp']) / 60.0)
                score_rain = 1.0 - (abs(rain - ideal['Rain']) / 3000.0)
                score_soil = 1.0 if soil_type == ideal['Soil'] else 0.5
                
                # Average score
                avg_score = (score_n + score_p + score_k + score_ph + score_temp + score_rain + score_soil) / 7.0
                match_percentage = round(max(0.0, min(1.0, avg_score)) * 100.0, 1)
                
                recommendations.append({
                    'crop': crop_name,
                    'match': match_percentage,
                    'soil': ideal['Soil'],
                    'ph': ideal['pH'],
                    'npk': f"N:{ideal['N']} P:{ideal['P']} K:{ideal['K']}"
                })
            
            # Sort by match score
            recommendations.sort(key=lambda x: x['match'], reverse=True)

    # 3. Market Price Comparison / price trends
    categories = Category.objects.all()
    trends = PriceTrend.objects.all().order_by('record_date')
    
    crop_trends = {}
    for trend in trends:
        if trend.crop_name not in crop_trends:
            crop_trends[trend.crop_name] = {'dates': [], 'prices': []}
        crop_trends[trend.crop_name]['dates'].append(trend.record_date.strftime("%Y-%m-%d"))
        crop_trends[trend.crop_name]['prices'].append(float(trend.avg_price))

    context = {
        'weather': weather_data,
        'form': form,
        'recommendations': recommendations,
        'crop_trends': crop_trends,
        'categories': categories,
    }
    return render(request, 'smart/smart_tools.html', context)


@login_required
def chat_list_view(request):
    user = request.user
    # Find all messages where user is sender or receiver
    messages_query = ChatMessage.objects.filter(Q(sender=user) | Q(receiver=user)).order_by('-created_at')
    
    # Extract unique contacts
    contacts = []
    contact_ids = set()
    for msg in messages_query:
        other_user = msg.receiver if msg.sender == user else msg.sender
        if other_user.id not in contact_ids:
            contact_ids.add(other_user.id)
            contacts.append({
                'user': other_user,
                'last_message': msg.message,
                'timestamp': msg.created_at,
                'unread': ChatMessage.objects.filter(sender=other_user, receiver=user, is_read=False).count()
            })
            
    # Suggest other users to chat with
    all_users = User.objects.exclude(id=user.id)
    if user.role == 'buyer':
        suggested_contacts = all_users.filter(role='farmer')
    else:
        suggested_contacts = all_users.filter(role='buyer')

    return render(request, 'marketplace/chat_list.html', {
        'contacts': contacts,
        'suggested_contacts': suggested_contacts[:6]
    })


@login_required
def chat_detail_view(request, username):
    user = request.user
    other_user = get_object_or_404(User, username=username)
    
    # Mark messages as read
    ChatMessage.objects.filter(sender=other_user, receiver=user, is_read=False).update(is_read=True)
    
    chat_messages = ChatMessage.objects.filter(
        (Q(sender=user) & Q(receiver=other_user)) |
        (Q(sender=other_user) & Q(receiver=user))
    ).order_by('created_at')
    
    crop_id = request.GET.get('crop_id')
    crop_context = None
    if crop_id:
        crop_context = Crop.objects.filter(id=crop_id).first()
        
    if request.method == 'POST':
        form = ChatMessageForm(request.POST)
        if form.is_valid():
            msg_text = form.cleaned_data['message']
            ChatMessage.objects.create(
                sender=user,
                receiver=other_user,
                crop=crop_context,
                message=msg_text
            )
            # Create system notification
            create_notification(
                other_user,
                f"New chat message from {user.username}: '{msg_text[:30]}...'",
                'system'
            )
            redirect_url = f"/chat/{username}/"
            if crop_id:
                redirect_url += f"?crop_id={crop_id}"
            return redirect(redirect_url)
    else:
        form = ChatMessageForm()
        
    return render(request, 'marketplace/chat_detail.html', {
        'other_user': other_user,
        'chat_messages': chat_messages,
        'form': form,
        'crop_context': crop_context
    })


def weather_dashboard_view(request):
    location = request.GET.get('location', 'Midwest Region')
    
    # Seed a local weather forecast simulator
    random.seed(hash(location))
    
    temp = random.randint(15, 38)
    humidity = random.randint(35, 95)
    wind_speed = random.randint(3, 30)
    rainfall = random.randint(0, 150)
    soil_moisture = random.randint(10, 80)
    uv_index = random.randint(1, 11)
    
    alerts = []
    if temp > 35:
        alerts.append("Heatwave Warning: Ensure high-frequency early morning irrigation to avoid crop stress.")
    elif temp < 5:
        alerts.append("Frost Warning: Cover delicate crops. Prevent surface freezing.")
    if wind_speed > 25:
        alerts.append("High Wind Warning: Secure greenhouse covers and prop up young tree crops.")
    if rainfall > 100:
        alerts.append("Flood Advisory: Inspect drainage trenches. Suspend fertilizer spraying.")
    elif rainfall == 0 and soil_moisture < 25:
        alerts.append("Drought Stress Warning: Drip irrigation should be activated immediately.")
        
    advisories = [
        {"crop": "Wheat", "status": "Good", "advice": "Ideal temperature for wheat grain development. No immediate action required."},
        {"crop": "Rice (Paddy)", "status": "Needs Attention", "advice": "Ensure paddy water level is maintained at 5-10cm. Soil moisture is critical."},
        {"crop": "Potatoes", "status": "Action Required", "advice": "Monitor soil moisture. Apply light irrigation if rainfall remains below 10mm."}
    ]
    
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    forecast = []
    for day in days:
        forecast.append({
            'day': day,
            'temp': temp + random.randint(-4, 4),
            'humidity': min(100, max(0, humidity + random.randint(-15, 15))),
            'condition': random.choice(['Sunny', 'Partly Cloudy', 'Overcast', 'Showers', 'Thunderstorms']),
            'precip': random.randint(0, 80)
        })
        
    context = {
        'location': location,
        'temp': temp,
        'humidity': humidity,
        'wind_speed': wind_speed,
        'rainfall': rainfall,
        'soil_moisture': soil_moisture,
        'uv_index': uv_index,
        'alerts': alerts,
        'advisories': advisories,
        'forecast': forecast
    }
    return render(request, 'smart/weather_dashboard.html', context)


def farmer_store_view(request, username):
    farmer_user = get_object_or_404(User, username=username, role='farmer')
    farmer = farmer_user.farmer_profile
    crops = farmer.crops.filter(is_approved=True)
    
    # Search and Filter inside farmer store
    q = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    if q:
        crops = crops.filter(models.Q(name__icontains=q) | models.Q(description__icontains=q))
    if category_id:
        crops = crops.filter(category_id=category_id)
        
    categories = Category.objects.filter(crops__farmer=farmer).distinct()
    
    # Aggregate reviews from farmer's crops
    reviews = Review.objects.filter(crop__farmer=farmer).order_by('-created_at')
    
    # Get similar farmers in same location/region
    related_farmers = Farmer.objects.exclude(id=farmer.id)[:3]
    
    has_purchased = False
    if request.user.is_authenticated and request.user.role == 'buyer':
        has_purchased = Order.objects.filter(
            buyer=request.user,
            farmer=farmer,
            status='Delivered'
        ).exists()
    
    context = {
        'farmer': farmer,
        'crops': crops,
        'categories': categories,
        'reviews': reviews,
        'related_farmers': related_farmers,
        'q': q,
        'selected_category': int(category_id) if category_id.isdigit() else '',
        'has_purchased': has_purchased,
    }
    return render(request, 'marketplace/farmer_store.html', context)


def market_prices_view(request):
    from marketplace.models import MarketInsight, PriceTrend
    
    insights = MarketInsight.objects.all()
    categories = Category.objects.all()
    
    # Search and Filter
    q = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    demand = request.GET.get('demand', '')
    sort = request.GET.get('sort', '')
    
    if q:
        insights = insights.filter(models.Q(crop_name__icontains=q) | models.Q(region__icontains=q))
    if category_id:
        insights = insights.filter(category_id=category_id)
    if demand:
        insights = insights.filter(demand_level=demand)
        
    if sort == 'price_desc':
        insights = insights.order_by('-current_price')
    elif sort == 'price_asc':
        insights = insights.order_by('current_price')
    elif sort == 'volume_desc':
        insights = insights.order_by('-volume_tonnes')
    else:
        insights = insights.order_by('-record_date')
        
    # Top gainers and losers
    gainers = MarketInsight.objects.filter(price_change_percent__gt=0).order_by('-price_change_percent')[:3]
    losers = MarketInsight.objects.filter(price_change_percent__lt=0).order_by('price_change_percent')[:3]
    
    # Price trends for Chart.js (Tomato, Wheat, Rice)
    trends = PriceTrend.objects.all().order_by('record_date')
    crop_trends = {}
    for trend in trends:
        if trend.crop_name not in crop_trends:
            crop_trends[trend.crop_name] = {'dates': [], 'prices': []}
        crop_trends[trend.crop_name]['dates'].append(trend.record_date.strftime("%Y-%m-%d"))
        crop_trends[trend.crop_name]['prices'].append(float(trend.avg_price))
        
    context = {
        'insights': insights,
        'categories': categories,
        'gainers': gainers,
        'losers': losers,
        'crop_trends': crop_trends,
        'q': q,
        'selected_category': int(category_id) if category_id.isdigit() else '',
        'selected_demand': demand,
        'selected_sort': sort,
    }
    return render(request, 'marketplace/market_prices.html', context)


# --- Custom Context Processor for Notifications & Cart Count ---
def global_vars(request):
    if request.user.is_authenticated:
        unread_notifications_count = Notification.objects.filter(user=request.user, is_read=False).count()
        cart_count = CartItem.objects.filter(user=request.user).aggregate(Sum('quantity'))['quantity__sum'] or 0
        return {
            'unread_notifications_count': unread_notifications_count,
            'cart_count': cart_count
        }
    return {
        'unread_notifications_count': 0,
        'cart_count': 0
    }


@login_required
def download_invoice_pdf_view(request, pk):
    order = get_object_or_404(Order, pk=pk)
    
    # Permission checks: only buyer, farmer of this order, or admin can download
    if not (request.user.is_admin() or request.user == order.buyer or (request.user.role == 'farmer' and request.user.farmer_profile == order.farmer)):
        messages.error(request, "Access Denied. You do not have permission to view this invoice.")
        return redirect('home')
        
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.graphics.shapes import Drawing, Rect, String
    from django.http import HttpResponse

    # Setup unique invoice number
    invoice_number = f"AC-{order.created_at.strftime('%Y%m%d')}-{order.id:04d}"
    invoice_date = order.created_at.strftime('%B %d, %Y')
    
    # Calculate costs
    tax = float(order.total_amount) * 0.05
    service_fee = 10.00
    grand_total = float(order.total_amount) + tax + service_fee
    
    # Create the buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4, 
        rightMargin=36, 
        leftMargin=36, 
        topMargin=36, 
        bottomMargin=36
    )
    
    # Define color scheme
    primary_green = colors.HexColor('#2E7D32')
    secondary_green = colors.HexColor('#4CAF50')
    accent_orange = colors.HexColor('#FF9800')
    neutral_dark = colors.HexColor('#333333')
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'InvoiceTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=primary_green,
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'InvoiceSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=neutral_dark,
        spaceAfter=15
    )
    header_right_style = ParagraphStyle(
        'HeaderRight',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#777777'),
        alignment=2 # Right aligned
    )
    meta_right_style = ParagraphStyle(
        'MetaRight',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=neutral_dark,
        alignment=2 # Right aligned
    )
    section_heading = ParagraphStyle(
        'SectionHeading',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=primary_green,
        spaceAfter=6
    )
    body_style = ParagraphStyle(
        'InvoiceBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=neutral_dark,
        leading=12
    )
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=neutral_dark
    )
    table_cell_bold_style = ParagraphStyle(
        'TableCellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=neutral_dark
    )
    
    elements = []
    
    # Header block
    header_left = [
        Paragraph("AgriConnect", title_style),
        Paragraph("Direct Farmer-to-Buyer Ecosystem", subtitle_style)
    ]
    header_right = [
        Paragraph("TAX INVOICE", header_right_style),
        Spacer(1, 8),
        Paragraph(f"<b>Invoice No:</b> {invoice_number}", meta_right_style),
        Paragraph(f"<b>Date:</b> {invoice_date}", meta_right_style),
        Paragraph(f"<b>Order ID:</b> #{order.id}", meta_right_style)
    ]
    
    header_table = Table([[header_left, header_right]], colWidths=[260, 260])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 15))
    
    # Separation Line
    elements.append(Table([[""]], colWidths=[520], rowHeights=[2], style=TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), primary_green),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ])))
    elements.append(Spacer(1, 15))
    
    # Addresses
    buyer_info = [
        Paragraph("<b>BUYER (BILL TO)</b>", section_heading),
        Spacer(1, 4),
        Paragraph(f"<b>Name:</b> {order.buyer.first_name or order.buyer.username}", body_style),
        Paragraph(f"<b>Phone:</b> {order.buyer.phone or 'N/A'}", body_style),
        Paragraph(f"<b>Address:</b> {order.shipping_address}", body_style),
    ]
    
    farmer_info = [
        Paragraph("<b>FARMER (SELLER)</b>", section_heading),
        Spacer(1, 4),
        Paragraph(f"<b>Farmer Name:</b> {order.farmer.user.first_name or order.farmer.user.username}", body_style),
        Paragraph(f"<b>Farm Name:</b> {order.farmer.farm_name or 'Local Farm'}", body_style),
        Paragraph(f"<b>Location:</b> {order.farmer.farm_location or 'Local Area'}", body_style),
        Paragraph(f"<b>Phone:</b> {order.farmer.user.phone or 'N/A'}", body_style),
    ]
    
    addr_table = Table([[buyer_info, farmer_info]], colWidths=[260, 260])
    addr_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(addr_table)
    elements.append(Spacer(1, 20))
    
    # Table of products
    table_data = [[
        Paragraph("Produce Description", table_header_style),
        Paragraph("Category", table_header_style),
        Paragraph("Unit Price", table_header_style),
        Paragraph("Quantity", table_header_style),
        Paragraph("Total Price", table_header_style)
    ]]
    
    for item in order.items.all():
        table_data.append([
            Paragraph(item.crop.name if item.crop else "Deleted Crop", table_cell_bold_style),
            Paragraph(item.crop.category.name if item.crop else "General", table_cell_style),
            Paragraph(f"₹{item.price_per_unit}", table_cell_style),
            Paragraph(f"{item.quantity} {item.crop.get_unit_display() if item.crop else 'kg'}", table_cell_style),
            Paragraph(f"₹{item.total_price():.2f}", table_cell_bold_style)
        ])
        
    prod_table = Table(table_data, colWidths=[150, 100, 90, 90, 90])
    prod_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_green),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 8),
        ('TOPPADDING', (0,1), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e8f5e9')),
    ]))
    elements.append(prod_table)
    elements.append(Spacer(1, 15))
    
    # Bottom block - Payment info on left, Calculations on right
    qr_drawing = Drawing(60, 60)
    qr_drawing.add(Rect(0, 0, 60, 60, fillColor=colors.white, strokeColor=primary_green, strokeWidth=1))
    qr_drawing.add(Rect(5, 40, 15, 15, fillColor=primary_green, strokeColor=None))
    qr_drawing.add(Rect(40, 40, 15, 15, fillColor=primary_green, strokeColor=None))
    qr_drawing.add(Rect(5, 5, 15, 15, fillColor=primary_green, strokeColor=None))
    qr_drawing.add(Rect(25, 25, 10, 10, fillColor=primary_green, strokeColor=None))
    qr_drawing.add(Rect(10, 25, 5, 5, fillColor=primary_green, strokeColor=None))
    qr_drawing.add(Rect(25, 10, 5, 5, fillColor=primary_green, strokeColor=None))
    
    verified_badge = Drawing(100, 20)
    verified_badge.add(Rect(0, 0, 100, 20, fillColor=colors.HexColor('#E8F5E9'), strokeColor=primary_green, strokeWidth=1, rx=5, ry=5))
    verified_badge.add(String(12, 6, "VERIFIED FARMER", fontName="Helvetica-Bold", fontSize=8, fillColor=primary_green))

    payment_info = [
        Paragraph("<b>PAYMENT INFORMATION</b>", section_heading),
        Spacer(1, 4),
        Paragraph("<b>Payment Method:</b> Cash On Delivery (COD)", body_style),
        Paragraph(f"<b>Payment Status:</b> {'Completed (Cash Collected)' if order.status == 'Delivered' else 'Pending Payment'}", body_style),
        Spacer(1, 8),
        verified_badge
    ]
    
    totals_table_data = [
        [Paragraph("Subtotal:", table_cell_style), Paragraph(f"₹{order.total_amount:.2f}", table_cell_bold_style)],
        [Paragraph("Taxes (5%):", table_cell_style), Paragraph(f"₹{tax:.2f}", table_cell_bold_style)],
        [Paragraph("Direct Service Fee:", table_cell_style), Paragraph(f"₹{service_fee:.2f}", table_cell_bold_style)],
        [Paragraph("Grand Total:", table_cell_bold_style), Paragraph(f"₹{grand_total:.2f}", ParagraphStyle('TotalGreen', parent=table_cell_bold_style, textColor=primary_green, fontSize=11))]
    ]
    
    totals_table = Table(totals_table_data, colWidths=[130, 100])
    totals_table.setStyle(TableStyle([
        ('LINEBELOW', (0,0), (-1,-2), 0.5, colors.HexColor('#eeeeee')),
        ('LINEBELOW', (0,-2), (-1,-1), 1, primary_green),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
    ]))
    
    left_column = [
        payment_info,
        Spacer(1, 10),
        Table([[qr_drawing, Paragraph("<font size=7.5 color='#777777'>Scan to verify invoice<br/>or check tracking status<br/>directly on AgriConnect</font>", body_style)]], colWidths=[70, 180], style=TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    ]
    
    summary_table = Table([[left_column, totals_table]], colWidths=[280, 240])
    summary_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Footer Notice
    footer_text = Paragraph(
        "<center><font color='#777777'>Thank you for supporting direct-trade sustainable commerce with AgriConnect.<br/>"
        "Need help? Reach customer support at <b>support@agriconnect.org</b> or call 1800-123-4567.</font></center>",
        body_style
    )
    elements.append(footer_text)
    
    # Build Document
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Invoice_{invoice_number}.pdf"'
    return response


@login_required
def notifications_list_view(request):
    notifications = Notification.objects.filter(user=request.user)
    
    # Filtering
    notif_type = request.GET.get('type', '')
    priority = request.GET.get('priority', '')
    
    if notif_type:
        notifications = notifications.filter(notification_type=notif_type)
    if priority:
        notifications = notifications.filter(priority=priority)
        
    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
    read_count = Notification.objects.filter(user=request.user, is_read=True).count()
    total_count = Notification.objects.filter(user=request.user).count()
    
    # Types stats
    type_counts = {}
    for choice in Notification.TYPE_CHOICES:
        type_counts[choice[0]] = Notification.objects.filter(user=request.user, notification_type=choice[0]).count()
        
    context = {
        'notifications': notifications,
        'unread_count': unread_count,
        'read_count': read_count,
        'total_count': total_count,
        'type_counts': type_counts,
        'selected_type': notif_type,
        'selected_priority': priority,
    }
    return render(request, 'marketplace/notifications.html', context)


@login_required
def notifications_mark_all_read_view(request):
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    messages.success(request, "All notifications marked as read.")
    return redirect('notifications_list')


@login_required
def notifications_mark_read_view(request, pk):
    notif = get_object_or_404(Notification, pk=pk, user=request.user)
    notif.is_read = True
    notif.save()
    
    if notif.link:
        return redirect(notif.link)
    return redirect('notifications_list')


@login_required
def notifications_delete_view(request, pk):
    notif = get_object_or_404(Notification, pk=pk, user=request.user)
    notif.delete()
    messages.success(request, "Notification deleted.")
    return redirect('notifications_list')


@login_required
def farmer_verification_submit_view(request):
    if not request.user.is_farmer():
        messages.error(request, "Access Denied. Farmers only.")
        return redirect('home')
        
    farmer = request.user.farmer_profile
    
    if request.method == 'POST':
        aadhaar = request.FILES.get('aadhaar_document')
        certificate = request.FILES.get('farm_certificate')
        
        if aadhaar:
            farmer.aadhaar_document = aadhaar
        if certificate:
            farmer.farm_certificate = certificate
            
        farmer.verification_status = 'review'
        farmer.save()
        
        # Notify farmer
        create_notification(
            request.user,
            "Verification documents submitted. Admin review is in progress.",
            'verification',
            'medium'
        )
        
        # Notify admins
        from accounts.models import User as AccountUser
        admins = AccountUser.objects.filter(role='admin')
        for admin in admins:
            create_notification(
                admin,
                f"New farmer verification request submitted by {request.user.username}.",
                'verification',
                'high',
                link=f"/admin/farmers/{farmer.id}/verify/"
            )
            
        messages.success(request, "Documents submitted successfully! Verification status is now Under Review.")
        return redirect('farmer_dashboard')
        
    context = {
        'farmer': farmer
    }
    return render(request, 'marketplace/farmer_verify.html', context)


@login_required
def admin_farmer_verify_view(request, pk):
    if not request.user.is_admin():
        messages.error(request, "Access Denied. Admins only.")
        return redirect('home')
        
    farmer = get_object_or_404(Farmer, pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action') # 'approve', 'reject', 'resubmit'
        notes = request.POST.get('admin_notes', '')
        
        if action == 'approve':
            farmer.verification_status = 'approved'
            farmer.certification_status = 'verified'
            farmer.trust_score = 95
            farmer.verification_date = timezone.now()
            farmer.verified_by = request.user
            msg_text = "Your verification request has been APPROVED! Verified Farmer badge activated with trust score 95/100."
            messages.success(request, f"Farmer {farmer.user.username} approved successfully!")
        elif action == 'reject':
            farmer.verification_status = 'rejected'
            farmer.certification_status = 'none'
            farmer.trust_score = 50
            msg_text = f"Your verification request was rejected. Reason: {notes}"
            messages.warning(request, f"Farmer {farmer.user.username} rejected.")
        else: # resubmit
            farmer.verification_status = 'resubmit'
            msg_text = f"Resubmission required for verification documents. Reason: {notes}"
            messages.info(request, f"Resubmission requested for farmer {farmer.user.username}.")
            
        farmer.admin_notes = notes
        farmer.save()
        
        # Notify farmer
        create_notification(
            farmer.user,
            msg_text,
            'verification',
            'high' if action != 'approve' else 'medium',
            link='/farmer/verify/'
        )
        
        return redirect('admin_dashboard')
        
    context = {
        'farmer': farmer
    }
    return render(request, 'marketplace/admin_farmer_verify.html', context)


@login_required
def wishlist_view(request):
    wishlist_items = Wishlist.objects.filter(user=request.user)
    
    # Search
    q = request.GET.get('q', '')
    if q:
        wishlist_items = wishlist_items.filter(
            Q(crop__name__icontains=q) | 
            Q(crop__farmer__user__username__icontains=q) | 
            Q(crop__category__name__icontains=q)
        )
        
    # Filters
    category_id = request.GET.get('category', '')
    if category_id:
        wishlist_items = wishlist_items.filter(crop__category_id=category_id)
        
    # Recommendations: crops from user's favorite farmers, or simply other organic/popular crops
    fav_farmers = FavoriteFarmer.objects.filter(buyer=request.user).values_list('farmer_id', flat=True)
    recommended_crops = Crop.objects.filter(is_approved=True, availability_status='available')
    if fav_farmers:
        recommended_crops = recommended_crops.filter(farmer_id__in=fav_farmers)
    else:
        recommended_crops = recommended_crops.order_by('?')
    recommended_crops = recommended_crops.exclude(id__in=wishlist_items.values_list('crop_id', flat=True))[:4]
    
    categories = Category.objects.all()
    
    context = {
        'wishlist_items': wishlist_items,
        'recommended_crops': recommended_crops,
        'categories': categories,
        'q': q,
        'selected_category': category_id,
        'wishlist_count': wishlist_items.count(),
        'fav_farmers_count': len(fav_farmers),
    }
    return render(request, 'marketplace/wishlist.html', context)


@login_required
def wishlist_toggle_view(request, crop_id):
    crop = get_object_or_404(Crop, id=crop_id)
    w_item = Wishlist.objects.filter(user=request.user, crop=crop)
    
    if w_item.exists():
        w_item.delete()
        status = 'removed'
        msg = f"'{crop.name}' removed from wishlist."
    else:
        Wishlist.objects.create(user=request.user, crop=crop)
        status = 'added'
        msg = f"'{crop.name}' added to your wishlist successfully."
        
    count = Wishlist.objects.filter(user=request.user).count()
    
    # Check if request is ajax/JSON request or form post
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        from django.http import JsonResponse
        return JsonResponse({'status': status, 'count': count, 'message': msg})
        
    messages.success(request, msg)
    return redirect(request.META.get('HTTP_REFERER', 'crop_list'))


@login_required
def wishlist_remove_view(request, pk):
    w_item = get_object_or_404(Wishlist, pk=pk, user=request.user)
    crop_name = w_item.crop.name
    w_item.delete()
    messages.success(request, f"'{crop_name}' removed from your wishlist.")
    return redirect('wishlist')


@login_required
def wishlist_move_to_cart_view(request, pk):
    w_item = get_object_or_404(Wishlist, pk=pk, user=request.user)
    crop = w_item.crop

    try:
        buyer_profile = request.user.buyer_profile
    except Buyer.DoesNotExist:
        buyer_profile = None

    if buyer_profile and not crop.is_eligible_for_delivery(buyer_profile):
        est_days = crop.estimated_delivery_days(buyer_profile)
        rem_days = crop.remaining_shelf_life_days
        messages.error(request, f"Cannot move to cart. This fresh crop has a remaining shelf life of {rem_days} days, but estimated delivery to {buyer_profile.city or 'your city'} takes {est_days} days. Order blocked to prevent spoilage.")
        return redirect('wishlist')
    
    # Create or update CartItem
    cart_item, created = CartItem.objects.get_or_create(
        user=request.user,
        crop=crop,
        defaults={'quantity': 1}
    )
    if not created:
        cart_item.quantity += 1
        cart_item.save()
        
    w_item.delete()
    messages.success(request, f"'{crop.name}' moved to shopping cart successfully.")
    return redirect('wishlist')


@login_required
def farmer_follow_view(request, farmer_id):
    farmer = get_object_or_404(Farmer, id=farmer_id)
    follow_rel = FavoriteFarmer.objects.filter(buyer=request.user, farmer=farmer)
    
    if follow_rel.exists():
        follow_rel.delete()
        messages.success(request, f"You have unfollowed {farmer.user.username}.")
    else:
        FavoriteFarmer.objects.create(buyer=request.user, farmer=farmer)
        messages.success(request, f"You are now following {farmer.user.username}!")
        
        # Notify farmer about new follower
        create_notification(
            farmer.user,
            f"Buyer {request.user.username} is now following your farm store updates!",
            'review',
            'medium'
        )
        
    return redirect(request.META.get('HTTP_REFERER', 'home'))


@login_required
def review_helpful_toggle_view(request, pk):
    review = get_object_or_404(Review, pk=pk)
    if request.user in review.helpful_votes.all():
        review.helpful_votes.remove(request.user)
        msg = "Removed helpful vote from review."
    else:
        review.helpful_votes.add(request.user)
        msg = "Marked review as helpful!"
    messages.success(request, msg)
    return redirect(request.META.get('HTTP_REFERER', 'crop_list'))


@login_required
def farmer_review_reply_view(request, pk):
    review = get_object_or_404(Review, pk=pk)
    if not request.user.is_farmer() or review.crop.farmer.user != request.user:
        messages.error(request, "Access Denied.")
        return redirect('home')
        
    if request.method == 'POST':
        reply_text = request.POST.get('reply', '')
        if reply_text:
            review.reply = reply_text
            review.reply_at = timezone.now()
            review.save()
            messages.success(request, "Your reply has been posted successfully.")
            # Notify buyer
            create_notification(
                review.buyer,
                f"Farmer {request.user.username} replied to your review on '{review.crop.name}'.",
                'review',
                'medium'
            )
    return redirect(request.META.get('HTTP_REFERER', 'crop_list'))


@login_required
def review_edit_view(request, pk):
    review = get_object_or_404(Review, pk=pk, buyer=request.user)
    if request.method == 'POST':
        form = ReviewForm(request.POST, request.FILES, instance=review)
        if form.is_valid():
            form.save()
            messages.success(request, "Your review has been updated.")
            return redirect('crop_detail', pk=review.crop.id)
    else:
        form = ReviewForm(instance=review)
    return render(request, 'marketplace/review_edit.html', {'form': form, 'review': review})


@login_required
def review_delete_view(request, pk):
    review = get_object_or_404(Review, pk=pk, buyer=request.user)
    crop_id = review.crop.id
    review.delete()
    messages.success(request, "Your review has been deleted.")
    return redirect('crop_detail', pk=crop_id)


@login_required
def farmer_rate_view(request, farmer_id):
    farmer = get_object_or_404(Farmer, id=farmer_id)
    if not request.user.role == 'buyer':
        messages.error(request, "Only buyers can rate farmers.")
        return redirect('home')
        
    from marketplace.forms import FarmerRatingForm
    from marketplace.models import FarmerRating
    
    # Check if they've bought from this farmer
    has_purchased = Order.objects.filter(
        buyer=request.user,
        farmer=farmer,
        status='Delivered'
    ).exists()
    
    if not has_purchased:
        messages.error(request, "You can only rate farmers from whom you have received a delivered order.")
        return redirect(request.META.get('HTTP_REFERER', 'home'))
        
    # Check if already rated
    existing_rating = FarmerRating.objects.filter(buyer=request.user, farmer=farmer).first()
    
    if request.method == 'POST':
        form = FarmerRatingForm(request.POST, instance=existing_rating)
        if form.is_valid():
            rating_obj = form.save(commit=False)
            rating_obj.buyer = request.user
            rating_obj.farmer = farmer
            rating_obj.save()
            
            # Update farmer store rating
            all_ratings = FarmerRating.objects.filter(farmer=farmer)
            avg_rating = all_ratings.aggregate(Avg('rating'))['rating__avg'] or 5.0
            farmer.store_rating = avg_rating
            farmer.save()
            
            messages.success(request, "Your rating for the farmer has been submitted successfully.")
            return redirect('farmer_store', username=farmer.user.username)
    else:
        form = FarmerRatingForm(instance=existing_rating)
        
    return render(request, 'marketplace/farmer_rate.html', {'form': form, 'farmer': farmer})


@login_required
def reports_dashboard_view(request):
    # Fetch report history for user
    reports = Report.objects.filter(generated_by=request.user).order_by('-created_at')
    
    # Calculate counters
    total_generated = reports.count()
    total_downloads = reports.aggregate(Sum('download_count'))['download_count__sum'] or 0
    active_schedules = reports.exclude(scheduled_interval__isnull=True).exclude(scheduled_interval='').count()
    
    # Fetch recent orders to populate charts (revenue analytics)
    # If farmer: sales from this farmer
    # If admin: all sales
    # If buyer: buyer purchases
    if request.user.role == 'admin':
        orders = Order.objects.all().order_by('-created_at')[:10]
        revenue_data = Order.objects.filter(status='Delivered').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    elif request.user.role == 'farmer':
        orders = Order.objects.filter(farmer=request.user.farmer_profile).order_by('-created_at')[:10]
        revenue_data = Order.objects.filter(farmer=request.user.farmer_profile, status='Delivered').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    else: # buyer
        orders = Order.objects.filter(buyer=request.user).order_by('-created_at')[:10]
        revenue_data = Order.objects.filter(buyer=request.user, status='Delivered').aggregate(Sum('total_amount'))['total_amount__sum'] or 0

    context = {
        'reports': reports,
        'total_generated': total_generated,
        'total_downloads': total_downloads,
        'active_schedules': active_schedules,
        'orders': orders,
        'revenue_data': float(revenue_data),
    }
    return render(request, 'marketplace/reports.html', context)


@login_required
def report_generate_view(request):
    if request.method == 'POST':
        report_type = request.POST.get('report_type', 'order')
        fmt = request.POST.get('format', 'csv')
        scheduled_interval = request.POST.get('scheduled_interval', '')
        
        name = f"{report_type.replace('_', ' ').title()} Report ({timezone.now().strftime('%Y-%m-%d')})"
        
        # Save to database
        Report.objects.create(
            name=name,
            report_type=report_type,
            generated_by=request.user,
            format=fmt,
            scheduled_interval=scheduled_interval if scheduled_interval else None
        )
        messages.success(request, f"New report '{name}' has been generated and added to your export log.")
        return redirect('reports_dashboard')
    return redirect('reports_dashboard')


@login_required
def report_download_view(request, pk):
    report = get_object_or_404(Report, pk=pk, generated_by=request.user)
    
    # Increment download count
    report.download_count += 1
    report.save()
    
    # Gather data based on report type
    headers = []
    data_rows = []
    
    if report.report_type == 'order':
        headers = ['Order ID', 'Buyer', 'Farmer', 'Total Amount (₹)', 'Payment Method', 'Status', 'Date']
        if request.user.role == 'admin':
            records = Order.objects.all()
        elif request.user.role == 'farmer':
            records = Order.objects.filter(farmer=request.user.farmer_profile)
        else:
            records = Order.objects.filter(buyer=request.user)
        for r in records:
            data_rows.append([str(r.id), r.buyer.username, r.farmer.farm_name, f"₹{r.total_amount}", r.payment_method, r.status, r.created_at.strftime('%Y-%m-%d %H:%M')])
            
    elif report.report_type == 'revenue':
        headers = ['Payment ID', 'Order ID', 'Amount (₹)', 'Method', 'Status', 'Date']
        if request.user.role == 'admin':
            records = Payment.objects.all()
        elif request.user.role == 'farmer':
            records = Payment.objects.filter(order__farmer=request.user.farmer_profile)
        else:
            records = Payment.objects.filter(order__buyer=request.user)
        for r in records:
            data_rows.append([str(r.id), str(r.order.id), f"₹{r.amount}", r.payment_method, r.status, r.created_at.strftime('%Y-%m-%d %H:%M')])
            
    elif report.report_type == 'product':
        headers = ['Product ID', 'Name', 'Category', 'Price/Unit (₹)', 'Stock Available', 'Unit', 'Approved']
        if request.user.role == 'admin':
            records = Crop.objects.all()
        elif request.user.role == 'farmer':
            records = Crop.objects.filter(farmer=request.user.farmer_profile)
        else:
            records = Crop.objects.filter(is_approved=True)
        for r in records:
            data_rows.append([str(r.id), r.name, r.category.name, f"₹{r.price_per_kg}", str(r.quantity_available), r.unit, "Yes" if r.is_approved else "No"])
            
    elif report.report_type == 'user':
        headers = ['User ID', 'Username', 'Email', 'Role', 'Date Joined']
        if request.user.role == 'admin':
            records = User.objects.all()
        else:
            records = [request.user]
        for r in records:
            data_rows.append([str(r.id), r.username, r.email, r.role, r.date_joined.strftime('%Y-%m-%d %H:%M')])
            
    else: # default fallback
        headers = ['Record ID', 'Created At']
        data_rows.append(['1', timezone.now().strftime('%Y-%m-%d %H:%M')])
        
    # Return formatted file based on requested format
    if report.format == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{report.name.replace(" ", "_")}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Marketplace Report"
        
        # Style Title
        title_font = Font(name='Arial', size=14, bold=True, color='2E7D32')
        ws['A1'] = "AgriConnect Marketplace Business Report"
        ws['A1'].font = title_font
        ws['A2'] = f"Report Name: {report.name}"
        ws['A3'] = f"Generated On: {report.created_at.strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Table Headers style
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            
        # Write data rows
        for row_idx, row_data in enumerate(data_rows, 6):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='left')
                
        # Auto-adjust column widths
        for col in ws.columns:
            vals = [cell.value for cell in col if cell.value is not None]
            max_len = max(len(str(v)) for v in vals) if vals else 10
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(response)
        return response

    elif report.format == 'pdf':
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{report.name.replace(" ", "_")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        story = []
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2E7D32'),
            spaceAfter=8
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#555555'),
            spaceAfter=20
        )
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10
        )
        header_cell_style = ParagraphStyle(
            'HeaderCellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            textColor=colors.white,
            fontName='Helvetica-Bold'
        )
        
        story.append(Paragraph("AgriConnect Business Intelligence Report", title_style))
        story.append(Paragraph(f"Report Name: {report.name} | Generated: {report.created_at.strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
        story.append(Spacer(1, 10))
        
        # Format table data
        table_data = []
        table_data.append([Paragraph(h, header_cell_style) for h in headers])
        for row in data_rows:
            table_data.append([Paragraph(str(val), cell_style) for val in row])
            
        col_count = len(headers)
        col_width = (letter[0] - 72) / col_count
        
        t = Table(table_data, colWidths=[col_width] * col_count)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2E7D32')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#F8FFF8'), colors.white]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
            ('BOTTOMPADDING', (0,1), (-1,-1), 6),
            ('TOPPADDING', (0,1), (-1,-1), 6),
        ]))
        
        story.append(t)
        doc.build(story)
        return response

    else: # Default CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{report.name.replace(" ", "_")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['AgriConnect Marketplace Business Report'])
        writer.writerow([f'Report Name: {report.name}', f'Generated On: {report.created_at}'])
        writer.writerow([])
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow(row)
        return response


@login_required
def report_delete_view(request, pk):
    report = get_object_or_404(Report, pk=pk, generated_by=request.user)
    name = report.name
    report.delete()
    messages.success(request, f"Report '{name}' deleted successfully.")
    return redirect('reports_dashboard')
